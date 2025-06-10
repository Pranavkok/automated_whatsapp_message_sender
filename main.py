import openpyxl
import pywhatkit
import time

# Note : Make sure you have the following libraries installed:
# pip install openpyxl
# pip install pywhatkit
# This script reads contacts from an Excel file 
# Make sure you have WhatsApp Web logged in on your default browser before running this script.
# Also, ensure that the headings  in the Excel file are in the first row 
# Make sure there is column named "contact" or "phone" or "mobile" or "number" or "phonenumber" or "mobilenumber" or "contactnumber" or "tel" thats stores the contact numbers.(its not case sensitive)
# Make sure there is column named "name" for send_msg_to_person

# whastapp message sending function
def send_whatsapp_message(contact, message):
    pywhatkit.sendwhatmsg_instantly(contact, message, wait_time=10, tab_close=True)

# path of the Excel file 
path = 

# Send message to all contacts in the Excel file 
def send_msg_to_all_contacts(path,message):
    # setting up the openpyxl
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    # finding the contact column
    flag = False
    for col in range(1,sheet_obj.max_column+1):
        cell = sheet_obj.cell(row=1, column=col)
        if(cell.value.lower().replace(" ", "") == ("contact" or "phone" or "mobile" or "number" or "phonenumber" or "mobilenumber" or "contactnumber" or "tel")):
            contact_col = col
            flag = True
            break
    
    # if Contact column doesn't exits 
    if not flag:
        print("Make sure that there is a Contact column in your excel file")
        exit()

    # iterate through all the rows of the contact column
    for row in range(2,sheet_obj.max_row+1):
        number = sheet_obj.cell(row=row , column=contact_col).value

        # checking whether the number starts with '+91'
        if not str(number).startswith('+91'):
            number = '+91' + str(number)

        # checking whether the number is valid
        length = len(number)
        if length == 13: 
            try:
                print(f"Sending message to : {number}")
                # sending the message
                send_whatsapp_message(number, message)
                print(f"Message sent to : {number}")
                time.sleep(5)
            except Exception as e:
                print(f"Failed to send message to {number}: {e}")

        else:
            print(f"Invalid number format for {number}, skipping...")
            continue 


# Send message to a specific person in the Excel file
def send_msg_to_person(path,name,message):
    # setting up the openpyxl
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    
    contact_col = None
    name_col = None
    Contact_flag = False
    Name_flag = False

    # finding the contact and Name column
    for col in range(1,sheet_obj.max_column+1):
        cell_value = sheet_obj.cell(row=1, column=col).value
        if not cell_value:
            continue
        cleaned = cell_value.lower().replace(" ", "")
        if cleaned in ["contact", "phone", "mobile", "number", "phonenumber", "mobilenumber", "contactnumber", "tel"]:
            contact_col = col
            Contact_flag = True
            
        elif(cleaned == "name"):
            name_col = col
            Name_flag = True
            
        if Contact_flag and Name_flag:
            break

    # if Contact and Name column doesn't exits 
    if not Contact_flag or not Name_flag:
        print("Make sure that there is a Contact column and Name column in your excel file")
        return

    # iterate through all the rows of the name column
    for row in range(2,sheet_obj.max_row+1):
        cell_name = sheet_obj.cell(row=row , column=name_col).value
        if(cell_name.lower().replace(" ","") == name.lower().replace(" ","")):
            number = sheet_obj.cell(row=row , column=contact_col).value
            # checking whether the number starts with '+91'
            if not str(number).startswith('+91'):
                number = '+91' + str(number)
            # checking whether the number is valid
            length = len(number)
            if length == 13: 
                try:
                    print(f"Sending message to : {number}")
                    # sending the message
                    send_whatsapp_message(number, message)
                    print(f"Message sent to : {number}")
                    time.sleep(5)
                    break
                except Exception as e:
                    print(f"Failed to send message to {number}: {e}")
                    
            else:
                print(f"Invalid number format for {number}, skipping...")


    

    
